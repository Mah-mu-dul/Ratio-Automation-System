from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import io
import math
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/health")
async def health_check():
    return {"status": "ok"}

def solve_knapsack(items, capacity):
    dp = [ {'val': -1, 'weight': 0, 'items': []} for _ in range(capacity + 1) ]
    dp[0] = {'val': 0, 'weight': 0, 'items': []}
    
    for item in items:
        w = item['weight']
        v = item['value']
        idx = item['index']
        for c in range(capacity, w - 1, -1):
            if dp[c - w]['val'] != -1:
                new_val = dp[c - w]['val'] + v
                new_weight = dp[c - w]['weight'] + w
                if new_val > dp[c]['val'] or (new_val == dp[c]['val'] and new_weight > dp[c]['weight']):
                    dp[c] = {
                        'val': new_val,
                        'weight': new_weight,
                        'items': dp[c - w]['items'] + [idx]
                    }
    
    best = max(dp, key=lambda x: (x['val'], x['weight']))
    return best['items']

def pack_plates(skus: List[Dict[str, Any]], plate_up: int, page_up: int, one_upc_plate_up: int, max_waste_pct: float):
    capacity = plate_up * page_up
    unassigned = skus.copy()
    plates = []
    
    while unassigned:
        candidate_runs = set()
        for item in unassigned:
            q = item['qty']
            for u in range(1, capacity + 1):
                candidate_runs.add(math.ceil(q / u))
        candidate_runs = sorted(list(candidate_runs))
        
        best_plate = None
        best_score = (-1, -1, -1)
        
        for R in candidate_runs:
            valid_items = []
            for i, item in enumerate(unassigned):
                q = item['qty']
                u = math.ceil(q / R)
                if u > capacity or u <= 0: continue
                waste = u * R - q
                waste_pct = (waste / q) * 100 if q > 0 else 0
                
                if waste_pct > max_waste_pct:
                    continue
                
                valid_items.append({
                    'index': i,
                    'weight': u,
                    'value': 1,
                    'waste': waste
                })
            
            if not valid_items:
                continue
                
            selected_indices = solve_knapsack(valid_items, capacity)
            if not selected_indices:
                continue
                
            selected_items = [v for v in valid_items if v['index'] in selected_indices]
            
            items_count = len(selected_items)
            utilization = sum(item['weight'] for item in selected_items)
            total_waste = sum(item['waste'] for item in selected_items)
            
            score = (items_count, utilization, -total_waste)
            
            if score > best_score:
                best_score = score
                best_plate = {
                    'R': R,
                    'items': selected_items
                }
        
        if best_plate:
            plate_items = []
            for item in best_plate['items']:
                idx = item['index']
                sku_info = unassigned[idx]
                plate_items.append({
                    'SIZE': sku_info['sku'],
                    'Plate Run': best_plate['R'],
                    'Repeat': float(item['weight']),
                    'Qty': sku_info['qty'],
                    'Waste': item['waste'],
                    'Rec#': float(sku_info['rec_num'])
                })
            plates.append(plate_items)
            
            assigned_indices = {item['index'] for item in best_plate['items']}
            unassigned = [item for i, item in enumerate(unassigned) if i not in assigned_indices]
        else:
            item = unassigned.pop(0)
            q = item['qty']
            u = one_upc_plate_up
            R = math.ceil(q / u)
            waste = u * R - q
            
            plates.append([{
                'SIZE': item['sku'],
                'Plate Run': R,
                'Repeat': float(u),
                'Qty': q,
                'Waste': waste,
                'Rec#': float(item['rec_num'])
            }])
            
    final_plates = []
    plate_id = 1
    for plate in plates:
        for item in plate:
            final_plates.append({
                'Plate#': float(plate_id),
                'SIZE': item['SIZE'],
                'Plate Run': item['Plate Run'],
                'Repeat': item['Repeat'],
                'Qty': item['Qty'],
                'Waste': item['Waste'],
                'Rec#': item['Rec#']
            })
        plate_id += 1
        
    return final_plates

@app.post("/api/calculate")
async def calculate_layout(
    file: UploadFile = File(...),
    plate_up: int = Form(10),
    page_up: int = Form(10),
    one_upc_plate_up: int = Form(10),
    waste_config: str = Form("4,5,6")
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file type")
    
    contents = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
        
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    sku_col = next((c for c in df.columns if 'sku' in c or 'item' in c or 'size' in c), None)
    qty_col = next((c for c in df.columns if 'qty' in c or 'quant' in c), None)
    
    if not sku_col or not qty_col:
        sku_col = df.columns[0]
        qty_col = df.columns[1]
        
    skus = []
    for idx, row in df.iterrows():
        sku_val = str(row[sku_col])
        try:
            qty_val = int(row[qty_col])
            if qty_val > 0:
                skus.append({'sku': sku_val, 'qty': qty_val, 'rec_num': idx + 1})
        except:
            pass

    waste_pcts = [float(x.strip()) for x in waste_config.split(',') if x.strip()]
    if not waste_pcts:
        waste_pcts = [5.0]

    results = {}
    for pct in waste_pcts:
        plates = pack_plates(skus, plate_up, page_up, one_upc_plate_up, pct)
        key = f"{int(pct)}%" if pct.is_integer() else f"{pct}%"
        results[key] = {
            'total_plates': len(set(p['Plate#'] for p in plates)),
            'plates': plates
        }
    
    return results

@app.post("/api/export")
async def export_layout(payload: Dict[str, Any]):
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    headers = ['Plate#', 'SIZE', 'Plate Run', 'Repeat', 'Qty', 'Waste', 'Rec#']
    
    for sheet_name, data in payload.items():
        ws = wb.create_sheet(title=sheet_name)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        for row_idx, p in enumerate(data['plates'], 2):
            ws.cell(row=row_idx, column=1, value=p.get('Plate#', ''))
            ws.cell(row=row_idx, column=2, value=p.get('SIZE', ''))
            ws.cell(row=row_idx, column=3, value=p.get('Plate Run', ''))
            ws.cell(row=row_idx, column=4, value=p.get('Repeat', ''))
            ws.cell(row=row_idx, column=5, value=p.get('Qty', ''))
            ws.cell(row=row_idx, column=6, value=p.get('Waste', ''))
            ws.cell(row=row_idx, column=7, value=p.get('Rec#', ''))
            
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Stepping")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": "attachment; filename=stepping.xls"}
    )
